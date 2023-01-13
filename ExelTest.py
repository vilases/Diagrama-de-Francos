#!/usr/bin/env python3
import os
import openpyxl as xl
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib


dat="2023/01"

def clear_screen():
    os.system("cls" if os.name== "nt" else "clear")

#Funcion para establecer si el dia es sabado o domingo. Solo en ese caso de True continua.
def datesod (dia):
    fecha=dat+"/"+str(dia)
# convert yyyy-mm-dd string to date object
    dt_object = datetime.strptime(fecha, "%Y/%m/%d").date()
    x=dt_object.weekday()
    if x==5:
        print("*** Sábado "+str(dt_object.day)+" ***")
    if x==6:
        print ("*** Domingo "+str(dt_object.day)    +" ***")
    if x==5 or x==6:
        return True

#Funcion que imprime cada dia del fin de semana con los respectivos inspectores que trabajan y sus posiciones.        
def finde(tabla,hoy):
    control=["Ctrl 0/7", "Ctrl 7/16" ,"Ctrl 16/24" , "Vol 0/7", "Vol 7/16", "Vol 16/24"]
    
    for i in range (int(hoy.day)+1,31):
        if datesod(tabla[i][0]):
            trab=[]
            for j in range(1,len(tabla[i])):
                if not tabla[i][j]==None:
                    if not tabla[i][j]=="Franco" :
                        print (tabla[0][j]," : ",tabla[i][j]) 
                        trab.append(tabla[i][j])
            for item in control:
                if not item in trab:
                    print (item, " SIN CUBRIR")
            print("-"*20) 

#Funcion que imprime por cada inspector los Francos y pociciones segun el dia. 
def francos(tabla):
    for j in range (1,len(tabla[0])):
        cont=0
        for i in range(0,31):
            if not tabla[i][j]==None:
                if tabla[i][j]=="Franco":
                    cont+=1
                if i==0:
                    print (tabla[i][j])
                    print ("\n")
                else:
                    print (i,"----",tabla[i][j]) 
        if cont<6:
            print(6-cont," franco a trabajar")
        print ("="*15)   
        print("\n")

#Funcion que toma los datos y los envia por mail a cada inspector. (falta crear una cuenta de mail y la base de datos con los mails de los inspectores"
def send_mail(msg, reciever):
   date=datetime.today()
   email= MIMEMultipart()
   email["From"]=""
   email["To"]=reciever
   msg["Subject"]="Francos ",date.year,"/",date.month
   password=""
   
#Aca se genera el objeto de la hoja de calculo y se guarda en una lista llamada "tabla"
wb=xl.Workbook()
wb=xl.load_workbook("hoja_de_prueba.xlsx")
hoja = wb.active
max_row=31
max_col=16
tabla=[]
for i in range(1,max_row+1):
    row=[]
    for j in range(1,max_col+1):
        cell=hoja.cell(row=i,column=j)
        row.append(cell.value)
    tabla.append(row)


hoy=datetime.today() 
opcion=""
opcion= input ("""Elija opcion:
1) Imprimir lista de francos.
2) Imprimir fin de semana.
3) Salir
""")

while opcion!="3":
    opcion= input ("""Elija opcion:
1) Imprimir lista de francos.
2) Imprimir fin de semana.
3) Salir
""")
    if opcion=="1":
        clear_screen()
        francos(tabla)
    elif opcion=="2":
        clear_screen()    
        finde(tabla,hoy)
    elif opcion=="3":
        clear_screen()
        print ("Hasta la próxima")
    else:
        print ("Opcion invalida")
     


#msg="Mensaje de Prueba<br>Prueba de salto de linea"
#send_mail(msg)

